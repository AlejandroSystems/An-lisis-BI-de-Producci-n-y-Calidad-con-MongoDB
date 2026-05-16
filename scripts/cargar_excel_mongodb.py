# cargar_excel_mongodb.py
# Objetivo: leer Excel con openpyxl y cargar datos a MongoDB sin usar pandas.
# Incluye lectura de precios por kg desde la hoja PRECIOS_KG
# para calcular pérdida económica estimada por merma.

from openpyxl import load_workbook
from pymongo import MongoClient, UpdateOne
from datetime import datetime


# ============================
# 1. CONFIGURACIÓN GENERAL
# ============================

RUTA_EXCEL = "data/reporte_erp_produccion.xlsm"

HOJA_BASE_DATOS = "BASE_DATOS"
HOJA_PRECIOS = "PRECIOS_KG"

MONGO_URI = "mongodb://localhost:27017/"
NOMBRE_BD = "bi_produccion_calidad"
NOMBRE_COLECCION = "reporte_erp_produccion"


# ============================
# 2. CONEXIÓN A MONGODB
# ============================

cliente = MongoClient(MONGO_URI)
db = cliente[NOMBRE_BD]
coleccion = db[NOMBRE_COLECCION]

# Crear índice único para evitar duplicados
coleccion.create_index("id_registro", unique=True)


# ============================
# 3. FUNCIONES AUXILIARES
# ============================

def obtener_valor(fila_dict, campo):
    """
    Devuelve el valor de un campo desde el diccionario de fila.
    Si no existe, devuelve None.
    """
    return fila_dict.get(campo)


def convertir_numero(valor):
    """
    Convierte valores numéricos de Excel a float.
    Si no puede convertir, devuelve 0.
    """
    try:
        if valor is None or valor == "":
            return 0
        return float(valor)
    except:
        return 0


def normalizar_texto(valor):
    """
    Normaliza textos para comparar nombres de productos.
    Evita problemas por mayúsculas, minúsculas o espacios.
    """
    if valor is None:
        return ""
    return str(valor).strip().lower()


def leer_encabezados(worksheet):
    """
    Lee los encabezados de la primera fila de una hoja Excel.
    """
    encabezados = []

    for celda in worksheet[1]:
        if celda.value is not None:
            encabezado = str(celda.value).strip().lower()
            encabezados.append(encabezado)
        else:
            encabezados.append(None)

    return encabezados


# ============================
# 4. LECTURA DEL EXCEL
# ============================

print("Leyendo archivo Excel...")

wb = load_workbook(RUTA_EXCEL, data_only=True)

ws_base = wb[HOJA_BASE_DATOS]
ws_precios = wb[HOJA_PRECIOS]


# ============================
# 5. LECTURA DE PRECIOS POR KG
# ============================

print("Leyendo hoja PRECIOS_KG...")

encabezados_precios = leer_encabezados(ws_precios)

precios_kg = {}

for fila in ws_precios.iter_rows(min_row=2, values_only=True):

    fila_precio = {}

    for i, valor in enumerate(fila):
        if i < len(encabezados_precios) and encabezados_precios[i] is not None:
            fila_precio[encabezados_precios[i]] = valor

    producto = fila_precio.get("producto")
    precio_kg = convertir_numero(fila_precio.get("precio_kg"))

    if producto is not None and str(producto).strip() != "":
        producto_normalizado = normalizar_texto(producto)
        precios_kg[producto_normalizado] = precio_kg

print(f"Productos con precio mapeado: {len(precios_kg)}")
print(precios_kg)


# ============================
# 6. LECTURA DE BASE_DATOS
# ============================

encabezados = leer_encabezados(ws_base)

print("Encabezados detectados en BASE_DATOS:")
print(encabezados)


# ============================
# 7. TRANSFORMACIÓN A DOCUMENTOS JSON
# ============================

documentos = []
productos_sin_precio = set()

for fila in ws_base.iter_rows(min_row=2, values_only=True):

    fila_dict = {}

    for i, valor in enumerate(fila):
        if i < len(encabezados) and encabezados[i] is not None:
            fila_dict[encabezados[i]] = valor

    id_registro = obtener_valor(fila_dict, "id_registro")

    # Evitar filas vacías
    if id_registro is None or str(id_registro).strip() == "":
        continue

    producto = obtener_valor(fila_dict, "producto")
    producto_normalizado = normalizar_texto(producto)

    kg_producidos = convertir_numero(obtener_valor(fila_dict, "kg_producidos"))
    kg_merma = convertir_numero(obtener_valor(fila_dict, "kg_merma"))
    porcentaje_merma = convertir_numero(obtener_valor(fila_dict, "porcentaje_merma"))

    # Buscar precio por kg según producto
    precio_kg = precios_kg.get(producto_normalizado, 0)

    if precio_kg == 0:
        productos_sin_precio.add(str(producto))

    # Cálculo económico
    perdida_estimada = kg_merma * precio_kg

    documento = {
        "id_registro": id_registro,
        "fecha": obtener_valor(fila_dict, "fecha"),
        "sede": obtener_valor(fila_dict, "sede"),
        "area": obtener_valor(fila_dict, "area"),
        "turno": obtener_valor(fila_dict, "turno"),
        "producto": producto,
        "lote": obtener_valor(fila_dict, "lote"),

        "produccion": {
            "kg_producidos": kg_producidos,
            "kg_merma": kg_merma,
            "porcentaje_merma": porcentaje_merma
        },

        "economico": {
            "precio_kg": precio_kg,
            "perdida_estimada": perdida_estimada
        },

        "calidad": {
            "estado_calidad": obtener_valor(fila_dict, "estado_calidad"),
            "tipo_incidencia": obtener_valor(fila_dict, "tipo_incidencia"),
            "criticidad": obtener_valor(fila_dict, "criticidad"),
            "accion_correctiva": obtener_valor(fila_dict, "accion_correctiva"),
            "requiere_seguimiento": obtener_valor(fila_dict, "requiere_seguimiento")
        },

        "responsable": {
            "nombre": obtener_valor(fila_dict, "responsable"),
            "cargo": obtener_valor(fila_dict, "cargo"),
            "supervisor": obtener_valor(fila_dict, "supervisor")
        },

        "observacion": obtener_valor(fila_dict, "observacion"),
        "fecha_registro": obtener_valor(fila_dict, "fecha_registro"),
        "estado_registro": obtener_valor(fila_dict, "estado_registro"),
        "fecha_carga_mongodb": datetime.now()
    }

    documentos.append(documento)


print(f"Registros preparados para MongoDB: {len(documentos)}")

if productos_sin_precio:
    print("Advertencia: existen productos sin precio mapeado:")
    for producto in productos_sin_precio:
        print(f"- {producto}")


# ============================
# 8. CARGA A MONGODB CON UPSERT
# ============================

print("Cargando datos a MongoDB...")

operaciones = []

for doc in documentos:
    operaciones.append(
        UpdateOne(
            {"id_registro": doc["id_registro"]},
            {"$set": doc},
            upsert=True
        )
    )

if operaciones:
    resultado = coleccion.bulk_write(operaciones)
    print("Carga finalizada correctamente.")
    print(f"Insertados: {resultado.upserted_count}")
    print(f"Actualizados: {resultado.modified_count}")
else:
    print("No hay datos para cargar.")

print("Proceso terminado.")